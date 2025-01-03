# export_manager.py
import pandas as pd
from typing import Dict, Optional
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

class ExportManager:
    def __init__(self, output_path: str):
        self.output_path = output_path

    def export_results(
        self,
        results: Dict,
        analyzer_type: str,
        include_visualization: bool = True
    ):
        """Export analysis results to Excel with formatting."""
        wb = Workbook()

        # Create Summary sheet
        self._create_summary_sheet(wb, results, analyzer_type)

        # Create detailed analysis sheets
        if "city_pivot" in results:
            self._create_pivot_sheet(wb, results["city_pivot"], "City_Analysis")
        if "class_pivot" in results:
            self._create_pivot_sheet(wb, results["class_pivot"], "Class_Analysis")

        # Save workbook
        wb.save(self.output_path)

    def _create_summary_sheet(
        self,
        wb: Workbook,
        results: Dict,
        analyzer_type: str
    ):
        """Create and format summary sheet."""
        ws = wb.active
        ws.title = f"{analyzer_type}_Summary"

        # Add title
        ws['A1'] = f"Market Analysis Results - {analyzer_type}"
        ws['A1'].font = Font(bold=True, size=14)

        # Add volume market share
        row = 3
        ws[f'A{row}'] = "Volume Market Share"
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        for brand, share in results["market_share"].items():
            ws[f'A{row}'] = brand
            ws[f'B{row}'] = f"{share:.1f}%"
            row += 1

        # Add value market share if available
        if "brand_values" in results:
            row += 1
            ws[f'A{row}'] = "Value Market Share"
            ws[f'A{row}'].font = Font(bold=True)

            row += 1
            for brand, value in results["brand_values"].items():
                ws[f'A{row}'] = brand
                ws[f'B{row}'] = f"{value:,.2f}"
                row += 1

    def _create_pivot_sheet(
        self,
        wb: Workbook,
        df: pd.DataFrame,
        sheet_name: str
    ):
        """Create and format pivot analysis sheet."""
        ws = wb.create_sheet(sheet_name)

        # Convert DataFrame to Excel
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="CCCCCC")